"""
환자 관리 API (스프레드시트 기반 파이프라인)

- CRUD + CSV 가져오기 + 퍼널 분석 + 동의 현황
- DB에 데이터 없으면 데모 데이터 반환
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func
from pydantic import BaseModel
from typing import Optional
from datetime import date, datetime
import json
import random
import logging
import uuid

from ..deps import get_db, get_current_active_user
from .service_guards import require_active_service
from ...models.user import User
from ...models.service_subscription import ServiceSubscription, ServiceType
from ...models.patient import Patient
from ...services.emr_import import (
    parse_file, auto_map, apply_mapping, MappedRow,
)

logger = logging.getLogger(__name__)
router = APIRouter()


# ============================================================
# Pydantic schemas
# ============================================================

class PatientCreate(BaseModel):
    name: str
    chart_no: Optional[str] = None
    phone: Optional[str] = None
    gender: Optional[str] = None
    birth_date: Optional[date] = None
    region: Optional[str] = None
    inflow_date: Optional[date] = None
    inflow_path: Optional[str] = None
    search_keywords: Optional[str] = None
    symptoms: Optional[str] = None
    diagnosis_name: Optional[str] = None
    consultation_summary: Optional[str] = None
    db_quality: Optional[str] = "MEDIUM"
    staff_assessment: Optional[str] = None
    appointment_date: Optional[datetime] = None
    appointment_path: Optional[str] = None
    inbound_status: Optional[str] = "PENDING"
    cancellation_reason: Optional[str] = None
    consultation_gap_analysis: Optional[str] = None
    manager_name: Optional[str] = None
    consent_examination: Optional[str] = "NOT_ASKED"
    consent_treatment: Optional[str] = "NOT_ASKED"
    partial_consent_reason: Optional[str] = None
    non_consent_reason: Optional[str] = None
    non_consent_root_cause: Optional[str] = None


class PatientUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    inbound_status: Optional[str] = None
    appointment_date: Optional[datetime] = None
    consent_examination: Optional[str] = None
    consent_treatment: Optional[str] = None
    manager_name: Optional[str] = None
    cancellation_reason: Optional[str] = None
    consultation_gap_analysis: Optional[str] = None
    partial_consent_reason: Optional[str] = None
    non_consent_reason: Optional[str] = None
    non_consent_root_cause: Optional[str] = None


# ============================================================
# Demo data
# ============================================================

DEMO_PATIENTS = [
    {"seq": 1, "chart": "P-001", "name": "김민수", "phone": "010-1111-2222", "gender": "M", "region": "강남구", "inflow_date": "2025-01-05", "inflow_path": "네이버 블로그", "keywords": "허리통증 내과", "symptoms": "만성 요통", "diagnosis": "요추 추간판 탈출증", "summary": "MRI 필요 상담", "quality": "HIGH", "assessment": "적극적 치료 의향", "status": "VISITED", "manager": "이실장", "consent_exam": "CONSENTED", "consent_treat": "CONSENTED"},
    {"seq": 2, "chart": "P-002", "name": "박지영", "phone": "010-2222-3333", "gender": "F", "region": "서초구", "inflow_date": "2025-01-08", "inflow_path": "네이버 광고", "keywords": "피부과 여드름", "symptoms": "여드름", "diagnosis": "심상성 여드름", "summary": "레이저 치료 상담", "quality": "MEDIUM", "assessment": "가격 비교 중", "status": "BOOKED", "manager": "김실장", "consent_exam": "CONSENTED", "consent_treat": "PARTIAL"},
    {"seq": 3, "chart": "P-003", "name": "이준호", "phone": "010-3333-4444", "gender": "M", "region": "송파구", "inflow_date": "2025-01-10", "inflow_path": "소개", "keywords": "", "symptoms": "건강검진", "diagnosis": "", "summary": "직장 건강검진", "quality": "HIGH", "assessment": "정기 고객 가능", "status": "VISITED", "manager": "이실장", "consent_exam": "CONSENTED", "consent_treat": "CONSENTED"},
    {"seq": 4, "chart": "P-004", "name": "최수정", "phone": "010-4444-5555", "gender": "F", "region": "강남구", "inflow_date": "2025-01-12", "inflow_path": "인스타그램", "keywords": "다이어트 한의원", "symptoms": "체중 관리", "diagnosis": "비만", "summary": "한약 치료 관심", "quality": "LOW", "assessment": "단순 문의", "status": "CANCELLED", "manager": "김실장", "consent_exam": "NOT_ASKED", "consent_treat": "NOT_ASKED"},
    {"seq": 5, "chart": "P-005", "name": "정태영", "phone": "010-5555-6666", "gender": "M", "region": "마포구", "inflow_date": "2025-01-15", "inflow_path": "구글 광고", "keywords": "내과 건강검진", "symptoms": "피로감", "diagnosis": "갑상선 기능 저하 의심", "summary": "혈액검사 권유", "quality": "MEDIUM", "assessment": "추가 검사 설득 필요", "status": "VISITED", "manager": "이실장", "consent_exam": "CONSENTED", "consent_treat": "REFUSED"},
    {"seq": 6, "chart": "P-006", "name": "한미래", "phone": "010-6666-7777", "gender": "F", "region": "강남구", "inflow_date": "2025-01-18", "inflow_path": "네이버 블로그", "keywords": "아토피 치료", "symptoms": "아토피 피부염", "diagnosis": "아토피 피부염", "summary": "면역 치료 상담", "quality": "HIGH", "assessment": "장기 치료 동의 가능", "status": "BOOKED", "manager": "김실장", "consent_exam": "CONSENTED", "consent_treat": "CONSENTED"},
    {"seq": 7, "chart": "P-007", "name": "윤성민", "phone": "010-7777-8888", "gender": "M", "region": "용산구", "inflow_date": "2025-01-20", "inflow_path": "오프라인 전단", "keywords": "", "symptoms": "두통", "diagnosis": "편두통", "summary": "진통제 처방 원함", "quality": "LOW", "assessment": "1회성 방문 예상", "status": "VISITED", "manager": "이실장", "consent_exam": "PARTIAL", "consent_treat": "REFUSED"},
    {"seq": 8, "chart": "P-008", "name": "서하늘", "phone": "010-8888-9999", "gender": "F", "region": "강동구", "inflow_date": "2025-01-22", "inflow_path": "카카오톡", "keywords": "감기 병원", "symptoms": "기침, 콧물", "diagnosis": "급성 상기도 감염", "summary": "일반 감기", "quality": "MEDIUM", "assessment": "재방문 가능성 낮음", "status": "VISITED", "manager": "김실장", "consent_exam": "CONSENTED", "consent_treat": "NOT_ASKED"},
    {"seq": 9, "chart": "P-009", "name": "조현우", "phone": "010-9999-0000", "gender": "M", "region": "서초구", "inflow_date": "2025-01-25", "inflow_path": "네이버 광고", "keywords": "위내시경 비용", "symptoms": "소화불량", "diagnosis": "기능성 소화불량", "summary": "내시경 검사 필요", "quality": "HIGH", "assessment": "검사 동의 확보 중", "status": "HELD", "manager": "이실장", "consent_exam": "PARTIAL", "consent_treat": "NOT_ASKED"},
    {"seq": 10, "chart": "P-010", "name": "임서연", "phone": "010-0000-1111", "gender": "F", "region": "강남구", "inflow_date": "2025-01-28", "inflow_path": "소개", "keywords": "", "symptoms": "고혈압", "diagnosis": "본태성 고혈압", "summary": "투약 시작 상담", "quality": "HIGH", "assessment": "장기 관리 환자", "status": "VISITED", "manager": "김실장", "consent_exam": "CONSENTED", "consent_treat": "CONSENTED"},
    {"seq": 11, "chart": "P-011", "name": "배동건", "phone": "010-1234-0001", "gender": "M", "region": "강남구", "inflow_date": "2025-02-01", "inflow_path": "네이버 블로그", "keywords": "당뇨 내과", "symptoms": "다음다갈", "diagnosis": "제2형 당뇨", "summary": "혈당 관리 상담", "quality": "HIGH", "assessment": "장기 관리 의향", "status": "VISITED", "manager": "이실장", "consent_exam": "CONSENTED", "consent_treat": "CONSENTED"},
    {"seq": 12, "chart": "P-012", "name": "노은비", "phone": "010-1234-0002", "gender": "F", "region": "마포구", "inflow_date": "2025-02-03", "inflow_path": "인스타그램", "keywords": "피부 레이저", "symptoms": "기미", "diagnosis": "기미", "summary": "레이저토닝 관심", "quality": "MEDIUM", "assessment": "가격 민감", "status": "PENDING", "manager": "김실장", "consent_exam": "NOT_ASKED", "consent_treat": "NOT_ASKED"},
]

INFLOW_PATH_COLORS = {
    "네이버 블로그": "emerald",
    "네이버 광고": "blue",
    "구글 광고": "red",
    "인스타그램": "purple",
    "카카오톡": "amber",
    "소개": "cyan",
    "오프라인 전단": "orange",
}


def _build_demo_patients() -> list[dict]:
    results = []
    for p in DEMO_PATIENTS:
        results.append({
            "id": f"demo-{p['seq']:03d}",
            "seq_no": p["seq"],
            "chart_no": p["chart"],
            "name": p["name"],
            "phone": p["phone"],
            "gender": p["gender"],
            "region": p["region"],
            "inflow_date": p["inflow_date"],
            "inflow_path": p["inflow_path"],
            "inflow_path_color": INFLOW_PATH_COLORS.get(p["inflow_path"], "gray"),
            "search_keywords": p["keywords"],
            "symptoms": p["symptoms"],
            "diagnosis_name": p["diagnosis"],
            "consultation_summary": p["summary"],
            "db_quality": p["quality"],
            "staff_assessment": p["assessment"],
            "inbound_status": p["status"],
            "manager_name": p["manager"],
            "consent_examination": p["consent_exam"],
            "consent_treatment": p["consent_treat"],
            "is_demo": True,
        })
    return results


# ============================================================
# Endpoints
# ============================================================

@router.get("")
@router.get("/", include_in_schema=False)
async def list_patients(
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    search: Optional[str] = None,
    status: Optional[str] = None,
    manager: Optional[str] = None,
    inflow_path: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """환자 목록 (pagination, search, filter). DB 우선, 비어있으면 데모 폴백."""
    # 실 DB 조회
    q_db = select(Patient).where(Patient.user_id == current_user.id)
    if search:
        like = f"%{search}%"
        q_db = q_db.where(or_(
            Patient.name.ilike(like),
            Patient.phone.ilike(like),
            Patient.chart_no.ilike(like),
        ))
    if status:
        q_db = q_db.where(Patient.inbound_status == status)
    if manager:
        q_db = q_db.where(Patient.manager_name == manager)
    if inflow_path:
        q_db = q_db.where(Patient.inflow_path == inflow_path)

    total_db = (await db.execute(
        select(func.count()).select_from(q_db.subquery())
    )).scalar() or 0

    if total_db > 0:
        rows = (await db.execute(
            q_db.order_by(Patient.created_at.desc())
                .offset((page - 1) * size).limit(size)
        )).scalars().all()
        items = [{
            "id": str(p.id),
            "chart_no": p.chart_no,
            "name": p.name,
            "phone": p.phone,
            "gender": p.gender,
            "birth_date": p.birth_date.isoformat() if p.birth_date else None,
            "inflow_path": p.inflow_path,
            "inbound_status": p.inbound_status.value if p.inbound_status else None,
            "manager_name": p.manager_name,
            "appointment_date": p.appointment_date.isoformat() if p.appointment_date else None,
        } for p in rows]
        return {"items": items, "total": total_db, "page": page, "size": size, "is_demo": False}

    # DB 비어있을 때만 데모 폴백
    patients = _build_demo_patients()
    if search:
        ql = search.lower()
        patients = [p for p in patients if ql in p["name"].lower() or ql in (p["phone"] or "") or ql in (p["chart_no"] or "").lower()]
    if status:
        patients = [p for p in patients if p["inbound_status"] == status]
    if manager:
        patients = [p for p in patients if p["manager_name"] == manager]
    if inflow_path:
        patients = [p for p in patients if p["inflow_path"] == inflow_path]

    total = len(patients)
    start = (page - 1) * size
    items = patients[start:start + size]
    return {"items": items, "total": total, "page": page, "size": size, "is_demo": True}


@router.get("/funnel/summary")
async def funnel_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """파이프라인 KPI"""
    patients = _build_demo_patients()
    total = len(patients)
    booked = len([p for p in patients if p["inbound_status"] in ["BOOKED", "VISITED"]])
    visited = len([p for p in patients if p["inbound_status"] == "VISITED"])
    cancelled = len([p for p in patients if p["inbound_status"] == "CANCELLED"])
    consented = len([p for p in patients if p["consent_treatment"] == "CONSENTED"])

    return {
        "total_inflow": total,
        "booking_rate": round(booked / total * 100, 1) if total else 0,
        "visit_rate": round(visited / total * 100, 1) if total else 0,
        "cancellation_rate": round(cancelled / total * 100, 1) if total else 0,
        "consent_rate": round(consented / total * 100, 1) if total else 0,
        "is_demo": True,
    }


@router.get("/funnel/stage-counts")
async def funnel_stage_counts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """단계별 환자 수"""
    patients = _build_demo_patients()
    stages = {
        "PENDING": {"label": "유입(대기)", "count": 0, "color": "gray"},
        "BOOKED": {"label": "예약완료", "count": 0, "color": "blue"},
        "HELD": {"label": "보류", "count": 0, "color": "amber"},
        "CANCELLED": {"label": "취소/이탈", "count": 0, "color": "red"},
        "VISITED": {"label": "내원완료", "count": 0, "color": "emerald"},
    }
    for p in patients:
        st = p["inbound_status"]
        if st in stages:
            stages[st]["count"] += 1

    return {"stages": list(stages.values()), "total": len(patients), "is_demo": True}


@router.get("/funnel/inflow-path")
async def funnel_inflow_path(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """유입경로별 전환율"""
    patients = _build_demo_patients()
    paths: dict[str, dict] = {}
    for p in patients:
        path = p["inflow_path"]
        if path not in paths:
            paths[path] = {"path": path, "color": p["inflow_path_color"], "total": 0, "booked": 0, "visited": 0, "consented": 0}
        paths[path]["total"] += 1
        if p["inbound_status"] in ["BOOKED", "VISITED"]:
            paths[path]["booked"] += 1
        if p["inbound_status"] == "VISITED":
            paths[path]["visited"] += 1
        if p["consent_treatment"] == "CONSENTED":
            paths[path]["consented"] += 1

    result = []
    for path_data in paths.values():
        t = path_data["total"]
        result.append({
            **path_data,
            "booking_rate": round(path_data["booked"] / t * 100, 1) if t else 0,
            "visit_rate": round(path_data["visited"] / t * 100, 1) if t else 0,
            "consent_rate": round(path_data["consented"] / t * 100, 1) if t else 0,
        })
    result.sort(key=lambda x: x["total"], reverse=True)
    return {"paths": result, "is_demo": True}


@router.get("/consent/dashboard")
async def consent_dashboard(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """동의 현황 대시보드"""
    patients = _build_demo_patients()
    visited = [p for p in patients if p["inbound_status"] == "VISITED"]
    total_visited = len(visited)

    # 검사 동의
    exam_consented = len([p for p in visited if p["consent_examination"] == "CONSENTED"])
    exam_partial = len([p for p in visited if p["consent_examination"] == "PARTIAL"])
    exam_refused = len([p for p in visited if p["consent_examination"] == "REFUSED"])

    # 치료 동의
    treat_consented = len([p for p in visited if p["consent_treatment"] == "CONSENTED"])
    treat_partial = len([p for p in visited if p["consent_treatment"] == "PARTIAL"])
    treat_refused = len([p for p in visited if p["consent_treatment"] == "REFUSED"])

    # 담당실장별 동의율
    manager_stats: dict[str, dict] = {}
    for p in visited:
        mgr = p["manager_name"] or "미배정"
        if mgr not in manager_stats:
            manager_stats[mgr] = {"manager": mgr, "total": 0, "consented": 0}
        manager_stats[mgr]["total"] += 1
        if p["consent_treatment"] == "CONSENTED":
            manager_stats[mgr]["consented"] += 1

    managers = []
    for ms in manager_stats.values():
        managers.append({
            **ms,
            "consent_rate": round(ms["consented"] / ms["total"] * 100, 1) if ms["total"] else 0,
        })

    # 미동의 사유 TOP5
    non_consent_reasons = [
        {"reason": "비용 부담", "count": 3},
        {"reason": "다른 병원 비교 후 결정", "count": 2},
        {"reason": "시간 부족", "count": 2},
        {"reason": "치료 필요성 미인식", "count": 1},
        {"reason": "가족 상의 필요", "count": 1},
    ]

    return {
        "total_visited": total_visited,
        "examination": {
            "consented": exam_consented,
            "partial": exam_partial,
            "refused": exam_refused,
            "rate": round(exam_consented / total_visited * 100, 1) if total_visited else 0,
        },
        "treatment": {
            "consented": treat_consented,
            "partial": treat_partial,
            "refused": treat_refused,
            "rate": round(treat_consented / total_visited * 100, 1) if total_visited else 0,
        },
        "by_manager": managers,
        "non_consent_reasons": non_consent_reasons,
        "is_demo": True,
    }


@router.get("/{patient_id}")
async def get_patient(
    patient_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """환자 상세 — DB 우선, 폴백으로 데모."""
    # UUID 형식이면 DB 조회
    try:
        from uuid import UUID
        pid = UUID(patient_id)
        row = (await db.execute(
            select(Patient).where(and_(
                Patient.id == pid, Patient.user_id == current_user.id,
            ))
        )).scalar_one_or_none()
        if row:
            return {
                "id": str(row.id),
                "chart_no": row.chart_no,
                "name": row.name,
                "phone": row.phone,
                "gender": row.gender,
                "birth_date": row.birth_date.isoformat() if row.birth_date else None,
                "region": row.region,
                "inflow_date": row.inflow_date.isoformat() if row.inflow_date else None,
                "inflow_path": row.inflow_path,
                "search_keywords": row.search_keywords,
                "symptoms": row.symptoms,
                "diagnosis_name": row.diagnosis_name,
                "consultation_summary": row.consultation_summary,
                "db_quality": row.db_quality.value if row.db_quality else None,
                "staff_assessment": row.staff_assessment,
                "appointment_date": row.appointment_date.isoformat() if row.appointment_date else None,
                "appointment_path": row.appointment_path,
                "inbound_status": row.inbound_status.value if row.inbound_status else None,
                "cancellation_reason": row.cancellation_reason,
                "consultation_gap_analysis": row.consultation_gap_analysis,
                "manager_name": row.manager_name,
                "consent_examination": row.consent_examination.value if row.consent_examination else None,
                "consent_treatment": row.consent_treatment.value if row.consent_treatment else None,
                "is_demo": False,
            }
    except (ValueError, TypeError):
        pass

    # 폴백: 데모
    patients = _build_demo_patients()
    patient = next((p for p in patients if p["id"] == patient_id), None)
    if not patient:
        if patients:
            patient = patients[0]
        else:
            raise HTTPException(status_code=404, detail="Not found")
    return {**patient, "is_demo": True}


@router.post("")
@router.post("/", include_in_schema=False)
async def create_patient(
    payload: PatientCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """신규 환자 등록 — 슬래시 유무 모두 받음 (Vercel rewrite 정규화 대응)."""
    patient = Patient(user_id=current_user.id, **payload.model_dump())
    db.add(patient)
    await db.commit()
    await db.refresh(patient)
    return {"id": str(patient.id), "message": "등록 완료"}


@router.put("/{patient_id}")
async def update_patient(
    patient_id: str,
    payload: PatientUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    result = await db.execute(
        select(Patient).where(and_(Patient.id == patient_id, Patient.user_id == current_user.id))
    )
    patient = result.scalar_one_or_none()
    if not patient:
        raise HTTPException(status_code=404, detail="Not found")
    for key, val in payload.model_dump(exclude_unset=True).items():
        setattr(patient, key, val)
    patient.updated_at = datetime.utcnow()
    await db.commit()
    return {"message": "수정 완료"}


@router.delete("/{patient_id}")
async def delete_patient(
    patient_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    result = await db.execute(
        select(Patient).where(and_(Patient.id == patient_id, Patient.user_id == current_user.id))
    )
    patient = result.scalar_one_or_none()
    if not patient:
        raise HTTPException(status_code=404, detail="Not found")
    await db.delete(patient)
    await db.commit()
    return {"message": "삭제 완료"}


# ============================================================
# 임포트 — 어떤 CSV/엑셀이든 자동 매핑 + 정규화
# ============================================================

@router.get("/import/template.xlsx")
async def import_template():
    """표준 임포트 템플릿 — 한국 EMR 호환 컬럼명 + 안내 행."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "환자목록"

    headers = [
        "차트번호", "이름", "성별", "생년월일", "전화번호", "지역",
        "유입일", "유입경로", "검색키워드", "주증상", "진단명",
        "상담요약", "DB등급", "실무자판단",
        "예약일", "예약경로", "내원상태", "취소사유",
        "담당실장", "검사동의", "치료동의",
    ]
    examples = [
        "A-001", "김환자", "남", "1985-03-15", "010-1234-5678", "서울 강남구",
        "2026-05-10", "네이버 광고", "허리통증 강남내과", "만성 요통", "요추 추간판 탈출증",
        "MRI 권유", "상", "적극 치료 의향",
        "2026-05-15 10:00", "전화", "예약", "",
        "이실장", "동의", "동의",
    ]
    note_row = [
        "필수: 이름 + (전화 또는 차트번호) — 둘 중 하나는 있어야 등록됩니다.",
        "", "M/F 또는 남/여 OK", "주민번호 앞자리도 자동 인식", "010 없어도 11자리 OK", "",
        "다양한 형식 OK (2026.05.10, 2026/5/10, 2026년 5월 10일)", "", "", "", "",
        "", "상/중/하 또는 H/M/L", "",
        "", "", "예약/대기/내원/취소 등", "",
        "", "동의/예/Y/거부/N 등", "동의/예/Y/거부/N 등",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    note_font = Font(italic=True, color="6B7280", size=9)
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for col, v in enumerate(examples, start=1):
        ws.cell(row=2, column=col, value=v)

    for col, v in enumerate(note_row, start=1):
        cell = ws.cell(row=3, column=col, value=v)
        cell.font = note_font

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = 18

    ws.cell(row=5, column=1, value="※ 알림톡 수신 동의는 정통망법/PIPA에 따라 사람이 직접 받은 동의만 인정합니다.")
    ws.cell(row=5, column=1).font = Font(italic=True, color="DC2626", size=9)
    ws.cell(row=6, column=1, value="※ 임포트 후 알림톡 동의는 모두 '미확인'으로 들어가며, 별도 화면에서 일괄 수정 가능합니다.")
    ws.cell(row=6, column=1).font = Font(italic=True, color="DC2626", size=9)
    ws.cell(row=7, column=1, value="※ 위 컬럼명을 그대로 쓰지 않아도 됩니다 — 시스템이 자동으로 매핑합니다.")
    ws.cell(row=7, column=1).font = Font(italic=True, color="6B7280", size=9)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="medimatch_patients_template.xlsx"'},
    )




# 베타 안전 한도 (2000명 가드와 정합)
_IMPORT_MAX_ROWS_PER_FILE = 5000   # 파서 자체 한도
_IMPORT_MAX_ROWS_PER_BATCH = 2000  # 한 번에 커밋 가능한 최대
_IMPORT_MAX_FILE_BYTES = 20 * 1024 * 1024  # 20 MB
_ALLOWED_EXTS = (".csv", ".tsv", ".txt", ".xlsx", ".xls", ".xlsm")


def _ext_ok(filename: str) -> bool:
    n = (filename or "").lower()
    return any(n.endswith(e) for e in _ALLOWED_EXTS)


async def _read_upload(file: UploadFile) -> bytes:
    raw = await file.read()
    if len(raw) > _IMPORT_MAX_FILE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"파일이 너무 큽니다 (최대 {_IMPORT_MAX_FILE_BYTES // 1024 // 1024}MB)",
        )
    return raw


def _row_to_preview(row: MappedRow) -> dict:
    """직렬화 가능한 형태로 변환 (date/datetime → ISO 문자열)."""
    out = {}
    for k, v in row.fields.items():
        if isinstance(v, (date, datetime)):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return {
        "fields": out,
        "external_meta_keys": list(row.external_meta.keys()),
        "issues": row.issues,
        "valid": row.is_valid(),
    }


@router.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...),
    manual_mapping: Optional[str] = Form(default=None),  # JSON: {"원본헤더": "canonical"}
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """업로드 → 자동 매핑 결과 + 처음 20행 미리보기 (실제 저장 X).

    프런트는 이 결과로 매핑 화면을 그리고, 사용자가 수정한 매핑을 다음 호출에 전달.
    """
    if not _ext_ok(file.filename or ""):
        raise HTTPException(
            status_code=400,
            detail=f"지원하지 않는 형식입니다. 가능: {', '.join(_ALLOWED_EXTS)}",
        )
    raw = await _read_upload(file)

    parsed = parse_file(file.filename or "upload", raw, max_rows=_IMPORT_MAX_ROWS_PER_FILE)
    if not parsed.headers:
        raise HTTPException(
            status_code=400,
            detail=parsed.warnings[0] if parsed.warnings else "파일을 읽지 못했습니다.",
        )

    manual = None
    if manual_mapping:
        try:
            manual = json.loads(manual_mapping)
            if not isinstance(manual, dict):
                raise ValueError("dict 형식이어야 합니다.")
        except (json.JSONDecodeError, ValueError) as e:
            raise HTTPException(status_code=400, detail=f"manual_mapping JSON 오류: {e}")

    plan = auto_map(parsed.headers, parsed.rows[:200], manual_mapping=manual)
    mapped = apply_mapping(plan, parsed.rows[:20])

    return {
        "filename": file.filename,
        "encoding": parsed.encoding_used,
        "sheet": parsed.sheet_used,
        "warnings": parsed.warnings,
        "total_rows": len(parsed.rows),
        "preview_rows": [_row_to_preview(r) for r in mapped],
        "mapping": plan.mapping,
        "confidence": plan.confidence,
        "unmapped_headers": plan.unmapped_headers,
        "detected_emr": plan.detected_emr,
        "notes": plan.notes,
        "valid_count_in_preview": sum(1 for r in mapped if r.is_valid()),
    }


@router.post("/import")
async def import_patients(
    file: UploadFile = File(...),
    manual_mapping: Optional[str] = Form(default=None),
    source_emr: str = Form(default="manual_csv"),
    skip_duplicates: bool = Form(default=True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """CSV/엑셀 → 환자 일괄 등록.

    안전 가드:
    - 한 파일 최대 _IMPORT_MAX_ROWS_PER_BATCH 행
    - 알림톡 동의는 무조건 NOT_ASKED (CSV 신뢰 X — 정통망법/PIPA)
    - (user_id, source_emr, external_id) unique — 같은 출처 중복 자동 skip
    - 행 단위 정규화 실패는 skip하고 리포트, 트랜잭션은 통째 커밋
    """
    if not _ext_ok(file.filename or ""):
        raise HTTPException(
            status_code=400,
            detail=f"지원하지 않는 형식입니다. 가능: {', '.join(_ALLOWED_EXTS)}",
        )
    raw = await _read_upload(file)

    parsed = parse_file(file.filename or "upload", raw, max_rows=_IMPORT_MAX_ROWS_PER_FILE)
    if not parsed.headers:
        raise HTTPException(
            status_code=400,
            detail=parsed.warnings[0] if parsed.warnings else "파일을 읽지 못했습니다.",
        )
    if len(parsed.rows) > _IMPORT_MAX_ROWS_PER_BATCH:
        raise HTTPException(
            status_code=400,
            detail=f"한 번에 최대 {_IMPORT_MAX_ROWS_PER_BATCH}명까지 임포트할 수 있습니다. "
                   f"파일을 나눠서 올려주세요.",
        )

    manual = None
    if manual_mapping:
        try:
            manual = json.loads(manual_mapping)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"manual_mapping JSON 오류: {e}")

    plan = auto_map(parsed.headers, parsed.rows[:200], manual_mapping=manual)
    mapped_rows = apply_mapping(plan, parsed.rows)

    batch_id = uuid.uuid4()
    now = datetime.utcnow()

    # 같은 출처에서 이미 가져온 external_id 미리 조회 (중복 검출용)
    existing_ext_ids: set[str] = set()
    if skip_duplicates:
        ext_ids_in_file = [
            r.fields.get("external_id") for r in mapped_rows
            if r.fields.get("external_id")
        ]
        if ext_ids_in_file:
            res = await db.execute(
                select(Patient.external_id).where(and_(
                    Patient.user_id == current_user.id,
                    Patient.source_emr == source_emr,
                    Patient.external_id.in_(ext_ids_in_file),
                ))
            )
            existing_ext_ids = {row[0] for row in res.all()}

    inserted = 0
    skipped_invalid = 0
    skipped_duplicate = 0
    issues: list[dict] = []

    for idx, mr in enumerate(mapped_rows, start=2):  # 2부터 = 헤더 다음 줄번호
        if not mr.is_valid():
            skipped_invalid += 1
            if len(issues) < 50:
                issues.append({"row": idx, "kind": "invalid", "detail": mr.issues})
            continue

        ext_id = mr.fields.get("external_id")
        if skip_duplicates and ext_id and ext_id in existing_ext_ids:
            skipped_duplicate += 1
            if len(issues) < 50:
                issues.append({"row": idx, "kind": "duplicate", "external_id": ext_id})
            continue

        # external_id가 같은 파일 내 중복 → 두 번째부터 skip
        if ext_id and ext_id in existing_ext_ids:
            skipped_duplicate += 1
            continue

        patient = Patient(
            user_id=current_user.id,
            source_emr=source_emr,
            external_id=ext_id,
            external_meta=mr.external_meta or None,
            import_batch_id=batch_id,
            imported_at=now,
            **{k: v for k, v in mr.fields.items() if k != "external_id"},
        )
        db.add(patient)
        inserted += 1
        if ext_id:
            existing_ext_ids.add(ext_id)

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        logger.exception("환자 임포트 커밋 실패")
        raise HTTPException(status_code=500, detail=f"DB 저장 실패: {e}")

    return {
        "batch_id": str(batch_id),
        "imported_count": inserted,
        "skipped_invalid": skipped_invalid,
        "skipped_duplicate": skipped_duplicate,
        "total_in_file": len(parsed.rows),
        "source_emr": source_emr,
        "mapping": plan.mapping,
        "unmapped_headers": plan.unmapped_headers,
        "issues": issues,
        "warnings": parsed.warnings,
        "notes": plan.notes,
        "message": f"{inserted}명 임포트 완료 (중복 {skipped_duplicate}, 불완전 {skipped_invalid})",
    }


@router.post("/import/rollback/{batch_id}")
async def rollback_import_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    sub: ServiceSubscription = Depends(require_active_service(ServiceType.EMR)),
):
    """배치 단위 롤백 — soft delete (의료법 5년 보존). 이번 임포트가 잘못됐을 때 즉시 되돌림."""
    try:
        bid = uuid.UUID(batch_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="잘못된 batch_id")

    res = await db.execute(
        select(Patient).where(and_(
            Patient.user_id == current_user.id,
            Patient.import_batch_id == bid,
            Patient.deleted_at.is_(None),
        ))
    )
    targets = res.scalars().all()
    now = datetime.utcnow()
    for p in targets:
        p.deleted_at = now
    await db.commit()
    return {"rolled_back": len(targets), "batch_id": batch_id}
